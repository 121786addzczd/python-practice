"""
employee_vacation_excel_generator.py
このスクリプトは、2023年度7月～9月最終日までの休暇を管理するエクセルファイルを生成します。
コンソールで「python employee_vacation_excel_generator.py」と入力すると実行できます。
実行すると「2023夏季休暇スケジュール.xlsx」ファイルが実行したディレクトリに生成されます。

注意事項:
このスクリプトを実行するには、以下の要件が必要です
- python version 3.9以上であること
- openpyxlのライブラリがあること 「pip install openpyxl」
- termcolorのライブラリがあること 「pip install  termcolor」
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date, timedelta
from openpyxl.styles import Font
import sys
import os
from termcolor import colored


def create_workbook(filename: str) -> Workbook:
    """ワークブックを作成する関数"""
    workbook = Workbook()
    sheet = workbook.active
    return workbook, sheet


def set_column_widths(sheet: Worksheet) -> None:
    """列の幅を設定する関数"""
    # CからCPまでの列を5に設定
    for col in range(3, 95):  # 列番号3から95まで
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 5
    
    sheet.column_dimensions['B'].width = 9


def create_title_row(sheet: Worksheet) -> None:
    """タイトル行を作成する関数"""
    sheet['B3'] = 'メンバー'
    sheet['B3'].alignment = Alignment(horizontal='center')


def fill_member_names(sheet: Worksheet, members: list[str]) -> None:
    """メンバーの名前を記入する関数"""
    for i, member in enumerate(members):
        row = i + 4  # メンバー名の行は3行目ではなく4行目から開始
        sheet.cell(row=row, column=2, value=member)  # 列もB列に変更
        sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')


def create_date_and_member_rows(
    sheet: Worksheet,
    start_date: date,
    end_date: date,
    weekdays: list[str],
    holidays: list[date],
    members: list[str]
) -> None:
    """日付とメンバーの行を作成する関数"""
    current_date = start_date
    col = 3  # 列もB列から開始

    while current_date <= end_date:
        cell = sheet.cell(row=2, column=col)
        cell.value = current_date.strftime('%-m/%-d')
        cell.alignment = Alignment(horizontal='center')

        # 曜日のヘッダーを追加
        weekday = weekdays[current_date.weekday()]
        cell = sheet.cell(row=3, column=col)
        cell.value = weekday
        cell.alignment = Alignment(horizontal='center')
        
        # 土曜日と日曜日の列に背景色を設定
        if weekday == '土' or weekday == '日':
            column_letter = get_column_letter(col)
            for row in range(3, sheet.max_row + 1):
                cell = sheet[column_letter + str(row)]
                gray_fill = PatternFill(fill_type="solid", fgColor="808080")
                cell.fill = gray_fill

        # 祝日のセルに背景色を設定
        if current_date in holidays:
            column_letter = get_column_letter(col)
            for row in range(3, sheet.max_row + 1):
                cell = sheet[column_letter + str(row)]
                holiday_fill = PatternFill(fill_type="solid", fgColor="808080")
                cell.fill = holiday_fill

        for j in range(len(members)):
            row = j + 4
            cell = sheet.cell(row=row, column=col)
            cell.value = ''
            cell.alignment = Alignment(horizontal='center')

        current_date += timedelta(days=1)
        col += 1

    sheet['B3'].value = '（敬称略）'


def set_cell_borders(sheet: Worksheet) -> None:
    """セルに枠線を設定する関数"""
    thin_border = Side(style='thin')
    no_border = Side(style=None)

    # 表内のセルに枠線を設定
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # 表の外側の枠線を設定
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # ヘッダーの枠線を設定
    for cell in sheet[2]:
        cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # B2とB3の左側に線を設定
    sheet['B2'].border = Border(left=thin_border, right=no_border, top=thin_border, bottom=thin_border)
    sheet['B3'].border = Border(left=thin_border, right=no_border, top=thin_border, bottom=thin_border)

    # A列の枠線を削除
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.border = Border(left=no_border, right=no_border, top=no_border, bottom=no_border)


def set_background_colors(sheet: Worksheet) -> None:
    """背景色を設定する関数"""
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
        for cell in row:
            cell.fill = PatternFill(fill_type="solid", fgColor="B8CCE4")

    # 日付のヘッダー行を塗る
    for col in range(2, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        cell = sheet[column_letter + '2']
        cell.fill = PatternFill(fill_type="solid", fgColor="B8CCE4")
    
    # 曜日のヘッダー行塗る
    for row in sheet.iter_rows(min_row=3, min_col=3, max_row=3, max_col=94):
        for cell in row:
            cell.fill = PatternFill(fill_type="solid", fgColor="B8CCE4")


def set_cell_values(sheet: Worksheet) -> None:
    """セルにテキストを記載する関数"""
    sheet['C1'] = '＊夏季休暇は3日間で、自由取得です。休暇予定日に「休」と記載ください。'
    
    sheet['S1'] = '※夏季休暇3日間の取得予定のみ記載ください（有給取得予定日等は記載しない）'
    sheet['S1'].font = Font(color='FF0000')
    
    cells = ['A12', 'A17', 'A22']
    value = '要調整'

    for cell in cells:
        sheet[cell] = value


def add_count_formula_cells(sheet: Worksheet) -> None:
    """合計の数をカウントするセルを追加する関数"""
    for row in range(4, sheet.max_row + 1):
        formula = f'=COUNTIF(B{row}:CO{row},"休")'
        if row == 4:
            cell = sheet.cell(row=row, column=sheet.max_column + 1)
        else:
            cell = sheet.cell(row=row, column=sheet.max_column)
        cell.value = formula
        cell.alignment = Alignment(horizontal='center')
    
    sheet['CQ3'] = '休暇取得数計'
    cell = sheet['CQ3']
    cell.fill = PatternFill(fill_type="solid", fgColor="F2DCDB")
    sheet.column_dimensions['CQ'].width = 12


def save_workbook(workbook: Workbook, filename: str) -> None:
    """ワークブックを保存する関数"""
    workbook.save(filename)


def main():
    try:
        # 保存するxlsxファイル名
        xlsx_file_name = '2023夏季休暇スケジュール.xlsx' # （必要に応じて書き換えてください）
        # 曜日のリスト
        weekdays = ['月', '火', '水', '木', '金', '土', '日']

        # メンバーのリスト（必要に応じて書き換えてください）
        members = ['佐藤', '鈴木', '高橋', '田中', '渡辺',
                    '伊藤', '山本', '中村', '小林', '加藤', 
                    '吉田', '山田', '佐々木', '山口', '松本', 
                    '井上', '木村', '清水', '谷口', '斉藤', 
                    '小川', '坂本']

        # 祝日のリスト（祝日の情報に基づいて追加してください）
        holidays = [date(2023, 7, 17), date(2023, 8, 11), date(2023, 9, 18)]
        
        try:
            # ファイルが既に開かれているかチェック
            if os.path.isfile(xlsx_file_name) and not os.access(xlsx_file_name, os.W_OK):
                raise Exception(colored(f"ファイル '{xlsx_file_name}' が既に開かれています", 'red'))
            elif os.path.isfile(xlsx_file_name):
                raise Exception(colored(f"ファイル '{xlsx_file_name}' は既に存在します", 'red'))

            # 休暇管理用のワークブックを作成
            workbook, sheet = create_workbook(xlsx_file_name)

        except Exception as e:
            print(str(e))
            sys.exit()

        # 休暇管理用のワークブックを作成
        workbook, sheet = create_workbook(xlsx_file_name)
        sheet.title = 'APL_Sys1'  # シート名を "APL_Sys1" に設定（必要に応じて書き換えてください）
        sheet.freeze_panes = 'C4' # 縦横スクロール固定

        set_column_widths(sheet)
        create_title_row(sheet)
        fill_member_names(sheet, members)
        create_date_and_member_rows(sheet, date(2023, 7, 1), date(2023, 9, 30), weekdays, holidays, members)
        set_cell_borders(sheet)
        set_background_colors(sheet)
        set_cell_values(sheet)
        add_count_formula_cells(sheet)

        # ファイルに保存
        save_workbook(workbook, xlsx_file_name)
        print(colored(f"{xlsx_file_name}を作成しました。", 'green'))
    except Exception as e:
        print(colored(f"エラーが発生しました: {e}", 'red'))
        sys.exit(1)

if __name__ == '__main__':
    main()