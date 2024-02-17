"""
create_schedule_from_source_sheet.py

このPythonスクリプトは、指定したエクセルファイルからスケジュールデータを取得し、新しいシートに書き込む機能を提供します。

概要:
- ユーザーはエクセルファイルのパス、作成するシートの名前、元のデータを取得するシートの名前を指定します。
- スクリプトは指定されたエクセルファイルを読み込み、エラーチェックを行います。
- 指定されたシートが既に存在する場合は削除し、新しいシートを作成します。
- 元のデータを取得してスケジュールデータを新しいシートに書き込みます。
- 変更内容を元のエクセルファイルに保存します。

使用方法:
1. create_schedule_from_source_sheet.pyファイルを実行します。
2. プログラムは指示に従い、エクセルファイルのパス、作成するシートの名前、元のデータを取得するシートの名前を入力します。
3. プログラムが処理を実行し、エクセルファイルにスケジュールデータが作成されます。

注意事項:
- 指定したエクセルファイルが存在しない場合や、正常なExcelファイルではない場合はエラーが発生します。
- ファイルにアクセス権限がない場合や、その他の予期しないエラーが発生した場合もエラーメッセージが表示されます。

依存関係:
- openpyxlライブラリがインストールされている必要があります。
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet


def create_or_clear_sheet(workbook, sheet_name) -> None:
    """
    シートが存在する場合は削除し、新しいシートを作成します。

    Args:
        workbook (Workbook): ワークブック
        sheet_name (str): シートの名前
    """
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
    workbook.create_sheet(sheet_name, 0)

def write_schedule_data(sheet, source_sheet):
    """
    シートに休暇スケジュールデータを書き込みます。

    Args:
        sheet (Worksheet): 書き込み先のシート
        source_sheet (Worksheet): データを取得する元のシート
    """
    row_index = 1

    # A1セルに「メンバー」、B1セルに「取得日」を書き込む
    sheet.cell(row=row_index, column=1, value='メンバー')
    sheet.cell(row=row_index, column=2, value='取得日')
    row_index += 1
    
    sheet.merge_cells('B1:D1')  # B1からD1までのセルを結合

    # 元のシートからデータを取得して書き込む
    for row in range(4, 26):
        name = source_sheet['B' + str(row)].value
        if name is not None:
            sheet.cell(row=row_index, column=1, value=name)

            column_index = 2
            for column in range(3, 95):
                date = source_sheet.cell(row=2, column=column).value
                value = source_sheet.cell(row=row, column=column).value
                if date is not None and value is not None:
                    sheet.cell(row=row_index, column=column_index, value=date)
                    column_index += 1

            row_index += 1

def decorate_sheet(sheet: Worksheet) -> None:
    """
    シートを装飾します。

    Args:
        sheet (Worksheet): 装飾するシート
    """
    # ヘッダー行の装飾
    header_fill = PatternFill(start_color="4F6228", end_color="4F6228", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_font.name = "游ゴシック"
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # A列の背景色を設定
    blue_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.fill = blue_fill

    # データ行の装飾
    cell_alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = cell_alignment

    # 枠線の装飾
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # シート全体のセルに枠線を適用
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

def create_schedule_from_source_sheet(file_path: str, sheet_name: str, source_sheet_name: str) -> None:
    """
    指定したエクセルファイルに元のシートからスケジュールを作成します。

    Args:
        file_path (str): エクセルファイルのパス
        sheet_name (str): 作成するシートの名前
        source_sheet_name (str): 元のシートの名前

    Raises:
        FileNotFoundError: ファイルが見つからない場合に発生
        InvalidFileException: 正常なExcelファイルではない場合に発生
        PermissionError: ファイルへのアクセス権限がない場合に発生
        Exception: その他のエラーが発生した場合に発生
    """
    try:
        # ファイルを開く
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"エラー: {file_path} が見つかりません。")
    except InvalidFileException:
        raise InvalidFileException(f"エラー: {file_path} は正常なExcelファイルではありません。")

    # 元のシートが存在するかチェック
    if source_sheet_name not in workbook.sheetnames:
        raise ValueError(f"エラー: {source_sheet_name} は存在しないシートです。")

    # シートが存在する場合は削除し、新しいシートを作成
    create_or_clear_sheet(workbook, sheet_name)

    # 元のシートを取得
    source_sheet = workbook[source_sheet_name]

    # データを書き込む
    sheet = workbook[sheet_name]
    write_schedule_data(sheet, source_sheet)
    
    # 見やすいようにsheet装飾
    decorate_sheet(sheet)

    # エクセルファイルを保存
    workbook.save(file_path)
    print(f"ファイルの保存が完了しました。ファイル名: {file_path}、シート名: {sheet_name}")


def main() -> None:
    file_path = '2023夏季休暇スケジュール.xlsx'
    source_sheet_name = 'APL_Sys1' # 抽出元となるシート
    sheet_name = f"休暇リスト{source_sheet_name}"

    try:
        create_schedule_from_source_sheet(file_path, sheet_name, source_sheet_name)
    except PermissionError:
        print(f"エラー: {file_path} にアクセスできません。別のプログラムでファイルが開かれている可能性があります。")
    except FileNotFoundError as e:
        print(f"エラー: {e}")
    except InvalidFileException as e:
        print(f"エラー: {e}")
    except ValueError as e:
        print(f"エラー: {e}")
    except Exception as e:
        print(f"エラーが発生しました: {e}")


if __name__ == '__main__':
    main()